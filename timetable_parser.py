from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
import requests
import openpyxl
from typing import List
import utils

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.66 Safari/537.36'}

s = requests.Session()
s.headers.update(headers)


blocksize = 12
blockstart = 4

class TimetableDay:
    def __init__(self, day_name=None, week_num=None):
        self.day_name = day_name
        self.week_num = week_num
        self.time_class_dict = {}

    def __repr__(self) -> str:
        return """{}, {} week: {}""".format(self.day_name, self.week_num, self.time_class_dict)

class Session:
    def __init__(self, addition=None, exams=None) -> None:
        self.addition = addition
        self.exams = exams

    def __repr__(self) -> str:
        return "{} {}".format(self.addition, self.exams)

def is_merged(sheet, coord):
    cell = sheet[coord]
    if isinstance(cell, MergedCell):
        return True
    return False
    

def cell_value(sheet, coord):
    cell = sheet[coord]
    if not isinstance(cell, MergedCell):
        return cell.value

#   print('merged')
    for range in sheet.merged_cells.ranges:
        if coord in range:
            return range.start_cell.value

    raise AssertionError('Merged cell is not in any merge range!')

def parse_day(sheet: Worksheet, start_row_num) -> List:
    day = cell_value(sheet, 'A{}'.format(start_row_num)).upper().replace(' ', '')
    day = utils.day_to_num(day)
    tday_w1 = TimetableDay(day, 1)
    tday_w2 = TimetableDay(day, 2)
    time = ""
    # print(day)

    for row in range(start_row_num, start_row_num+blocksize):
        print_msg = ''
        for col in range(1, 3):
            if (col == 1):
                time = cell_value(sheet, 'B{}'.format(row))
                print_msg += time + ' '
            else:
                coord = 'C{}'.format(row)
                val_sec = sheet['D{}'.format(row)].value
                val = cell_value(sheet, coord)

                if val_sec != None:
                    val = '1 подгруппа: {} / 2 подгруппа: {}'.format(val, val_sec)
                
                val = str(val).replace('\n', ' ').replace('\xa0', ' ')
                if (row%2 == 0):
                    tday_w1.time_class_dict[time] = val
                else:
                    tday_w2.time_class_dict[time] = val
                print_msg += val
        # print(print_msg)
    d = [tday_w1, tday_w2]

    return d

def parse_regular_timetable(sheet: Worksheet)-> List[TimetableDay]:
    week = []
    cur_start_row = blockstart
    for _ in range(0,6):
        try:
            week_days = parse_day(sheet, cur_start_row)
            week.extend(week_days)
            cur_start_row += blocksize
        except Exception as e:
            print(e)
    
    return week

def parse_session_timetable(sheet: Worksheet)->List[Session]:
    ret = []
    addition = sheet.cell(2, 1).value
    exams = {}
    max_ses_rows = 40
    for row in range(4, max_ses_rows):
        try:
            date = str(sheet.cell(row, 1).value) + ' ' + str(sheet.cell(row, 2).value)
            data = sheet.cell(row, 3).value
            if data != None:
                # print(dt, data.replace('\n', ' '))
                date = date.split(" ")[1]
                exams[date] = data.replace('\n', ' ')[:len(data)//2]
        except Exception as e:
            print(e)
    ret.append(Session(addition, exams))
    return ret
            

def get_timetable_from_url(url: str, is_session: bool=None):
    #url = r"https://pstu.ru/files/file/Abitur/timetable/2021-2022%20Raspisanie%20ehkzamenov%20EHTF%20ASU%20-19-1b%20%28osennijj%20%20sessiya%29.xlsx"
    response = s.get(url)
    if response.status_code == 200:
        with open("rasp.xlsx", 'wb') as f:
            f.write(response.content)
    print(url)
    wb2 = openpyxl.load_workbook('rasp.xlsx')
    shname = wb2.sheetnames[0]
    ws = wb2[shname]

    if is_session:
        ret = parse_session_timetable(ws)
    else:
        ret = parse_regular_timetable(ws)
    wb2.close()
    
    return ret
